def wb_form_c01():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Wartości_do_przepisania'
    ws.cell(1, 1, value='index_formularza')
    ws.cell(1, 2, value='5')
    ws.cell(1, 3, value='6')
    ws.cell(1, 4, value='7')
    ws.cell(1, 5, value='8')
    ws.cell(1, 6, value='9')
    ws.cell(1, 7, value='10')
    ws.cell(1, 8, value='11')
    ws.cell(1, 9, value='12')
    ws.cell(2, 1, value='01')
    ws.cell(3, 1, value='02')
    ws.cell(4, 1, value='03')
    ws.cell(5, 1, value='04')

    filename = 'New/C-01.xlsx'
    wb.save(filename)
