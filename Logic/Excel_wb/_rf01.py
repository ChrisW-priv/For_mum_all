def wb_form_rf01():
    from openpyxl import Workbook

    wb = Workbook()

    for i in range(2, 4+1):
        wb.create_sheet(title=str(i))

    def create_content():
        for i in range(2, 4+1):
            ws = wb[str(i)]
            ws.cell(1, 1, value='id_nr')
            ws.cell(1, 2, value='value')

    create_content()

    filename = 'New/RF-01.xlsx'
    wb.save(filename)
